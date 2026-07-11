from __future__ import annotations

import hashlib
import json
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import (
    CommentAnalysisResult,
    CommentNetwork,
    CommentNetworkEdge,
    CommentNetworkNode,
    CommentSnapshot,
    ReportSnapshot,
    ScriptAnalysisResult,
    TranscriptSegment,
)


class FileStorage:
    def __init__(self, root: Path | str) -> None:
        self.root = Path(root).resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def write(self, relative_path: str, content: bytes) -> tuple[str, int, str]:
        path = (self.root / relative_path).resolve()
        if self.root not in path.parents:
            raise ValueError("invalid storage path")
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)
        return str(path), len(content), hashlib.sha256(content).hexdigest()


class HtmlReportRenderer:
    def __init__(self) -> None:
        template_dir = Path(__file__).with_name("templates")
        self.environment = Environment(
            loader=FileSystemLoader(template_dir),
            autoescape=select_autoescape(("html", "xml")),
        )

    def render(self, report: ReportSnapshot) -> str:
        return self.environment.get_template("report.html").render(report=report, payload=report.payload)


class ExcelExporter:
    def render(self, session: Session, report: ReportSnapshot) -> bytes:
        from io import BytesIO

        workbook = Workbook()
        summary = workbook.active
        summary.title = "summary"
        summary.append(["field", "value"])
        for key, value in _flatten(report.payload):
            summary.append([key, value])

        comments = workbook.create_sheet("comment_analysis")
        comments.append(["youtube_comment_id", "is_reply", "text", "status", "is_hate_speech", "categories", "reasoning"])
        for comment, result in session.execute(
            select(CommentSnapshot, CommentAnalysisResult)
            .join(CommentAnalysisResult, CommentAnalysisResult.comment_snapshot_id == CommentSnapshot.id)
            .where(CommentAnalysisResult.analysis_run_id == report.analysis_run_id)
        ):
            comments.append([comment.youtube_comment_id, comment.is_reply, comment.text_original, result.status, result.is_hate_speech, _json(result.categories), result.reasoning])

        scripts = workbook.create_sheet("script_analysis")
        scripts.append(["segment_index", "start_seconds", "end_seconds", "text", "status", "is_hate_speech", "categories", "reasoning"])
        for segment, result in session.execute(
            select(TranscriptSegment, ScriptAnalysisResult)
            .join(ScriptAnalysisResult, ScriptAnalysisResult.transcript_segment_id == TranscriptSegment.id)
            .where(ScriptAnalysisResult.analysis_run_id == report.analysis_run_id)
        ):
            scripts.append([segment.segment_index, float(segment.start_seconds or 0), float(segment.end_seconds or 0), segment.text, result.status, result.is_hate_speech, _json(result.categories), result.reasoning])

        nodes = workbook.create_sheet("network_nodes")
        nodes.append(["node_key", "label", "comment_count", "hate_speech_count", "hate_speech_ratio", "metrics"])
        network = session.scalar(select(CommentNetwork).where(CommentNetwork.analysis_run_id == report.analysis_run_id))
        if network:
            for node in session.scalars(select(CommentNetworkNode).where(CommentNetworkNode.network_id == network.id)):
                nodes.append([node.node_key, node.label, node.comment_count, node.hate_speech_count, float(node.hate_speech_ratio), _json(node.metrics)])

        edges = workbook.create_sheet("network_edges")
        edges.append(["source_node_key", "target_node_key", "edge_type", "weight", "is_hate_speech"])
        if network:
            for edge in session.scalars(select(CommentNetworkEdge).where(CommentNetworkEdge.network_id == network.id)):
                edges.append([edge.source_node_key, edge.target_node_key, edge.edge_type, float(edge.weight), edge.is_hate_speech])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


def _flatten(value, prefix=""):
    if isinstance(value, dict):
        for key, item in value.items():
            yield from _flatten(item, f"{prefix}.{key}" if prefix else key)
    else:
        yield prefix, _json(value) if isinstance(value, (list, dict)) else value


def _json(value) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)
