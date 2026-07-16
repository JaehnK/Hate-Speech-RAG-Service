from io import BytesIO

import httpx
import pytest
from openpyxl import load_workbook

from app.core.config import Settings
from app.db.base import Base
from app.db.models import (
    AnalysisRun,
    CommentAnalysisResult,
    CommentSnapshot,
    ScriptAnalysisResult,
    TranscriptSegment,
    TranscriptSnapshot,
    VideoMetadataSnapshot,
)
from app.db.repositories import AnalysisJobRepository
from app.main import create_app
from app.reporting.builder import ReportBuilder
from app.reporting.network import CommentNetworkBuilder
from app.reporting.renderers import ExcelExporter, HtmlReportRenderer


def test_network_report_html_and_excel_builders(tmp_path) -> None:
    app = _app(tmp_path)
    with app.state.session_factory.begin() as session:
        report = _seed_report(session)
        assert report.payload["comment_analysis_summary"]["hate_speech_count"] == 1
        assert report.payload["network_summary"]["node_count"] == 2
        assert report.payload["analysis_config"]["retriever_config"]["example_min_similarity"] == 0.4
        report.payload["representative_comments"][0]["text"] = "<script>alert(1)</script>"
        rendered = HtmlReportRenderer().render(report)
        assert "대표 사례" in rendered
        assert "<script>alert(1)</script>" not in rendered
        assert "&lt;script&gt;" in rendered
        workbook = load_workbook(BytesIO(ExcelExporter().render(session, report)))
        assert workbook.sheetnames == ["summary", "comment_analysis", "script_analysis", "network_nodes", "network_edges"]
        assert workbook["comment_analysis"].max_row == 3


@pytest.mark.asyncio
async def test_report_detail_export_and_admin_apis(tmp_path) -> None:
    app = _app(tmp_path)
    with app.state.session_factory.begin() as session:
        report = _seed_report(session)
        report_id = report.id
        job = session.get(AnalysisRun, report.analysis_run_id)
        repository = AnalysisJobRepository(session)
        steps = repository.list_steps(job.job_id)
        failed = next(step for step in steps if step.step_key == "collect_comments")
        failed.status = "failed"
        failed.error_code = "YOUTUBE_QUOTA_EXCEEDED"
        repository.get(job.job_id).status = "partial_success"
        job_id = job.job_id

    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        report_response = await client.get(f"/api/reports/{report_id}")
        comments = await client.get(f"/api/reports/{report_id}/comments", params={"is_hate_speech": True})
        scripts = await client.get(f"/api/reports/{report_id}/script-segments")
        network = await client.get(f"/api/reports/{report_id}/network")
        page = await client.get(f"/reports/{report_id}")
        assert report_response.status_code == 200
        assert len(comments.json()["items"]) == 1
        assert len(scripts.json()["items"]) == 1
        assert len(network.json()["edges"]) == 1
        assert "text/html" in page.headers["content-type"]

        for format in ("html", "xlsx"):
            created = await client.post(f"/api/reports/{report_id}/exports", json={"format": format})
            assert created.status_code == 202
            export_id = created.json()["export_id"]
            status_response = await client.get(f"/api/exports/{export_id}")
            download = await client.get(status_response.json()["download_url"])
            assert status_response.json()["status"] == "succeeded"
            assert download.status_code == 200
            assert download.content
        rejected = await client.post(f"/api/reports/{report_id}/exports", json={"format": "pdf"})
        assert rejected.status_code == 400

        unauthorized = await client.get("/api/admin/jobs")
        headers = {"X-Admin-Token": "test-admin-secret"}
        admin_jobs = await client.get("/api/admin/jobs", headers=headers)
        settings = await client.get("/api/admin/settings", headers=headers)
        retry = await client.post(f"/api/admin/jobs/{job_id}/retry", headers=headers)
        assert unauthorized.status_code == 401
        assert admin_jobs.status_code == 200
        assert "test-admin-secret" not in settings.text
        assert retry.status_code == 202
        assert retry.json()["status"] == "pending"


@pytest.mark.asyncio
async def test_report_comments_sort_hate_speech_by_likes(tmp_path) -> None:
    app = _app(tmp_path)
    with app.state.session_factory.begin() as session:
        report = _seed_report(session)
        run = session.get(AnalysisRun, report.analysis_run_id)
        popular = CommentSnapshot(
            job_id=run.job_id,
            youtube_video_id="abcdefghijk",
            youtube_comment_id="popular-hate",
            text_original="popular",
            like_count=20,
        )
        session.add(popular)
        session.flush()
        session.add(
            CommentAnalysisResult(
                analysis_run_id=run.id,
                comment_snapshot_id=popular.id,
                status="succeeded",
                is_hate_speech=True,
                categories=["identity"],
            )
        )
        report_id = report.id

    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        first = await client.get(
            f"/api/reports/{report_id}/comments",
            params={"is_hate_speech": True, "sort": "like_count", "limit": 1},
        )
        second = await client.get(
            f"/api/reports/{report_id}/comments",
            params={"is_hate_speech": True, "sort": "like_count", "limit": 1, "cursor": 1},
        )

    assert first.status_code == 200
    assert first.json()["items"][0]["youtube_comment_id"] == "popular-hate"
    assert first.json()["total"] == 2
    assert first.json()["next_cursor"] == 1
    assert second.json()["items"][0]["like_count"] == 5
    assert not second.json()["has_more"]


def _app(tmp_path):
    app = create_app(
        Settings(
            database_url=f"sqlite:///{tmp_path / 'reporting.db'}",
            report_storage_dir=str(tmp_path / "reports"),
            admin_token="test-admin-secret",
        )
    )
    Base.metadata.create_all(app.state.engine)
    return app


def _seed_report(session):
    job = AnalysisJobRepository(session).create("abcdefghijk", "abcdefghijk")
    session.add(VideoMetadataSnapshot(job_id=job.id, youtube_video_id="abcdefghijk", title="Test Video", channel_title="Test Channel", comment_count=2))
    parent = CommentSnapshot(
        job_id=job.id,
        youtube_video_id="abcdefghijk",
        youtube_comment_id="parent",
        author_display_name="Parent",
        author_channel_id="author-parent",
        text_original="normal",
        like_count=1,
    )
    session.add(parent)
    session.flush()
    reply = CommentSnapshot(
        job_id=job.id,
        youtube_video_id="abcdefghijk",
        youtube_comment_id="reply",
        parent_youtube_comment_id="parent",
        parent_comment_snapshot_id=parent.id,
        is_reply=True,
        reply_depth=1,
        author_display_name="Reply",
        author_channel_id="author-reply",
        text_original="hate reply",
        like_count=5,
    )
    transcript = TranscriptSnapshot(job_id=job.id, youtube_video_id="abcdefghijk", source_type="public_caption", status="succeeded")
    session.add_all([reply, transcript])
    session.flush()
    segment = TranscriptSegment(transcript_snapshot_id=transcript.id, segment_index=0, text="normal script")
    run = AnalysisRun(
        job_id=job.id,
        youtube_video_id="abcdefghijk",
        status="running",
        llm_provider="fake",
        llm_model="fake",
        embedding_provider="hash",
        embedding_model="hash",
        example_vector_collection="examples",
        definition_vector_collection="definitions",
        definition_corpus_version="test-v1",
        retriever_config={"example_min_similarity": 0.4},
        prompt_versions={"comment": "test-v1", "script": "test-v1"},
    )
    session.add_all([segment, run])
    session.flush()
    session.add_all(
        [
            CommentAnalysisResult(analysis_run_id=run.id, comment_snapshot_id=parent.id, status="succeeded", is_hate_speech=False, categories=["unclassified"]),
            CommentAnalysisResult(analysis_run_id=run.id, comment_snapshot_id=reply.id, status="succeeded", is_hate_speech=True, categories=["gender"], reasoning="targeted"),
            ScriptAnalysisResult(analysis_run_id=run.id, transcript_segment_id=segment.id, status="succeeded", is_hate_speech=False, categories=["unclassified"]),
        ]
    )
    session.flush()
    CommentNetworkBuilder().build(session, run)
    return ReportBuilder().build(session, run)
