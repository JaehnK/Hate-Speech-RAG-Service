from pathlib import Path

from openpyxl import Workbook

from experiments.io import read_jsonl
from experiments.prepare_legacy_gold import (
    SOURCE_COMMENTS,
    SOURCE_INTERCODER,
    SOURCE_SCRIPTS,
    build_review_queue,
    convert_workbook,
)


def test_convert_legacy_comments_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "comments.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "comment_id",
            "video_id",
            "author",
            "comment_text",
            "is_hate_speech",
            "categories",
        ]
    )
    sheet.append(["c1", "v1", "@user", "혐오 댓글", True, "['정체성']"])
    workbook.save(workbook_path)

    input_path = tmp_path / "inputs.jsonl"
    legacy_path = tmp_path / "legacy.jsonl"
    input_count, legacy_count = convert_workbook(workbook_path, SOURCE_COMMENTS, input_path, legacy_path)

    assert (input_count, legacy_count) == (1, 1)
    assert read_jsonl(input_path) == [{"item_id": "c1", "source_type": "comment", "text": "혐오 댓글"}]
    legacy = read_jsonl(legacy_path)[0]
    assert legacy["legacy_is_hate_speech"] is True
    assert legacy["legacy_categories"] == ["정체성"]
    assert legacy["text_hash"]


def test_convert_legacy_scripts_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "scripts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["video_id", "script_index", "input_text", "is_hate_speech", "categories"])
    sheet.append(["v1", 3, "자막 문장", False, "['미분류']"])
    workbook.save(workbook_path)

    input_path = tmp_path / "inputs.jsonl"
    legacy_path = tmp_path / "legacy.jsonl"
    convert_workbook(workbook_path, SOURCE_SCRIPTS, input_path, legacy_path)

    assert read_jsonl(input_path)[0] == {
        "item_id": "v1:script:3",
        "source_type": "script",
        "text": "자막 문장",
    }
    assert read_jsonl(legacy_path)[0]["legacy_is_hate_speech"] is False


def test_convert_intercoder_workbook_skips_codebook_row(tmp_path: Path) -> None:
    workbook_path = tmp_path / "intercoder.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "for_Intercoder"
    sheet.append(["영상id", "id", "comment_text", "FC"])
    sheet.append(["번호", "번호", "설명", "최종분류"])
    sheet.append(["v1", "v1_0", "실제 댓글", "오정보"])
    workbook.save(workbook_path)

    input_path = tmp_path / "inputs.jsonl"
    legacy_path = tmp_path / "legacy.jsonl"
    input_count, legacy_count = convert_workbook(workbook_path, SOURCE_INTERCODER, input_path, legacy_path)

    assert (input_count, legacy_count) == (1, 1)
    assert read_jsonl(input_path)[0]["text"] == "실제 댓글"
    assert read_jsonl(legacy_path)[0]["legacy_fc"] == "오정보"
    assert read_jsonl(legacy_path)[0]["legacy_is_hate_speech"] is None


def test_build_review_queue_marks_agreement_and_disagreement(tmp_path: Path) -> None:
    legacy_path = tmp_path / "legacy.jsonl"
    legacy_path.write_text(
        "\n".join(
            [
                '{"item_id":"1","source_type":"comment","legacy_is_hate_speech":true,"legacy_categories":["정체성"],"text_hash":"a"}',
                '{"item_id":"2","source_type":"comment","legacy_is_hate_speech":false,"legacy_categories":["미분류"],"text_hash":"b"}',
                '{"item_id":"3","source_type":"comment","legacy_is_hate_speech":null,"legacy_categories":[],"legacy_fc":"오정보","text_hash":"c"}',
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    current_path = tmp_path / "current.jsonl"
    current_path.write_text(
        "\n".join(
            [
                '{"item_id":"1","variant":"three_vector_rag","status":"succeeded","payload":{"is_hate_speech":true,"categories":["정체성"]}}',
                '{"item_id":"2","variant":"three_vector_rag","status":"succeeded","payload":{"is_hate_speech":true,"categories":["정체성"]}}',
                '{"item_id":"3","variant":"three_vector_rag","status":"succeeded","payload":{"is_hate_speech":false,"categories":["미분류"]}}',
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    queue_path = tmp_path / "queue.jsonl"

    queue = build_review_queue(legacy_path, current_path, queue_path)

    assert [row["review_state"] for row in queue] == ["auto_candidate", "needs_review", "needs_review"]
    assert queue[1]["review_reason"] == "legacy_current_binary_disagreement"
    assert queue[2]["review_reason"] == "legacy_has_no_hate_label"
    assert read_jsonl(queue_path) == queue
