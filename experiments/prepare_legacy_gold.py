from __future__ import annotations

import argparse
import ast
import hashlib
import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from experiments.io import read_jsonl


SOURCE_COMMENTS = "legacy-comments"
SOURCE_SCRIPTS = "legacy-scripts"
SOURCE_INTERCODER = "intercoder"


def convert_workbook(
    workbook_path: Path | str,
    source: str,
    input_output_path: Path | str,
    legacy_output_path: Path | str,
    limit: int | None = None,
) -> tuple[int, int]:
    rows = list(_iter_legacy_rows(Path(workbook_path), source, limit=limit))
    input_rows = [
        {
            "item_id": row["item_id"],
            "source_type": row["source_type"],
            "text": row["text"],
        }
        for row in rows
    ]
    legacy_rows = [
        {
            "item_id": row["item_id"],
            "source_type": row["source_type"],
            "text_hash": _stable_hash(row["text"]),
            "legacy_is_hate_speech": row.get("legacy_is_hate_speech"),
            "legacy_categories": row.get("legacy_categories", []),
            "legacy_fc": row.get("legacy_fc"),
            "provenance": row["provenance"],
        }
        for row in rows
    ]
    _write_jsonl(input_output_path, input_rows)
    _write_jsonl(legacy_output_path, legacy_rows)
    return len(input_rows), len(legacy_rows)


def build_review_queue(
    legacy_label_path: Path | str,
    current_results_path: Path | str,
    output_path: Path | str,
    judge_label_path: Path | str | None = None,
) -> list[dict[str, Any]]:
    legacy_rows = {str(row["item_id"]): row for row in read_jsonl(legacy_label_path)}
    current_rows = _latest_successful_results(read_jsonl(current_results_path))
    judge_rows = {str(row["item_id"]): row for row in read_jsonl(judge_label_path)} if judge_label_path else {}

    queue = []
    for item_id, legacy in sorted(legacy_rows.items()):
        current = current_rows.get(item_id)
        judge = judge_rows.get(item_id)
        decision = _decide_review_state(legacy, current, judge)
        queue.append(
            {
                "item_id": item_id,
                "source_type": legacy.get("source_type", "comment"),
                "review_state": decision["state"],
                "review_reason": decision["reason"],
                "legacy": _legacy_summary(legacy),
                "current": _result_summary(current),
                "judge": _result_summary(judge),
            }
        )
    _write_jsonl(output_path, queue)
    return queue


def _iter_legacy_rows(workbook_path: Path, source: str, limit: int | None = None) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["for_Intercoder"] if source == SOURCE_INTERCODER else workbook[workbook.sheetnames[0]]
    headers = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    count = 0
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values, strict=False))
        converted = _convert_row(row, source, workbook_path)
        if converted is None:
            continue
        yield converted
        count += 1
        if limit is not None and count >= limit:
            break


def _convert_row(row: dict[str, Any], source: str, workbook_path: Path) -> dict[str, Any] | None:
    if source == SOURCE_COMMENTS:
        text = _clean_text(row.get("comment_text"))
        item_id = _clean_text(row.get("comment_id"))
        if not text or not item_id:
            return None
        return {
            "item_id": item_id,
            "source_type": "comment",
            "text": text,
            "legacy_is_hate_speech": _bool_or_none(row.get("is_hate_speech")),
            "legacy_categories": _parse_categories(row.get("categories")),
            "legacy_fc": None,
            "provenance": {"source": source, "path": str(workbook_path), "video_id": row.get("video_id")},
        }
    if source == SOURCE_SCRIPTS:
        text = _clean_text(row.get("input_text"))
        video_id = _clean_text(row.get("video_id"))
        script_index = row.get("script_index")
        if not text or not video_id or script_index is None:
            return None
        return {
            "item_id": f"{video_id}:script:{script_index}",
            "source_type": "script",
            "text": text,
            "legacy_is_hate_speech": _bool_or_none(row.get("is_hate_speech")),
            "legacy_categories": _parse_categories(row.get("categories")),
            "legacy_fc": None,
            "provenance": {"source": source, "path": str(workbook_path), "video_id": video_id, "script_index": script_index},
        }
    if source == SOURCE_INTERCODER:
        text = _clean_text(row.get("comment_text"))
        item_id = _clean_text(row.get("id"))
        fc = _clean_text(row.get("FC"))
        if not text or not item_id or fc == "최종분류":
            return None
        return {
            "item_id": item_id,
            "source_type": "comment",
            "text": text,
            "legacy_is_hate_speech": None,
            "legacy_categories": [],
            "legacy_fc": fc or None,
            "provenance": {"source": source, "path": str(workbook_path), "video_id": row.get("영상id")},
        }
    raise ValueError(f"unsupported source: {source}")


def _latest_successful_results(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    latest = {}
    for row in rows:
        item_id = str(row.get("item_id"))
        if row.get("status") != "succeeded":
            latest.setdefault(item_id, row)
            continue
        current = latest.get(item_id)
        if current is None or int(row.get("repeat_index", 0)) >= int(current.get("repeat_index", 0)):
            latest[item_id] = row
    return latest


def _decide_review_state(
    legacy: dict[str, Any],
    current: dict[str, Any] | None,
    judge: dict[str, Any] | None = None,
) -> dict[str, str]:
    if current is None or current.get("status") != "succeeded" or not current.get("payload"):
        return {"state": "needs_review", "reason": "current_missing_or_failed"}
    legacy_hate = legacy.get("legacy_is_hate_speech")
    if legacy_hate is None:
        return {"state": "needs_review", "reason": "legacy_has_no_hate_label"}

    current_payload = current["payload"]
    current_hate = bool(current_payload.get("is_hate_speech"))
    if bool(legacy_hate) != current_hate:
        return {"state": "needs_review", "reason": "legacy_current_binary_disagreement"}

    legacy_categories = _normal_categories(legacy.get("legacy_categories", []))
    current_categories = _normal_categories(current_payload.get("categories", []))
    if legacy_categories and legacy_categories != current_categories:
        return {"state": "needs_review", "reason": "legacy_current_category_disagreement"}

    if judge is not None:
        judge_payload = judge.get("payload", judge)
        judge_hate = judge_payload.get("is_hate_speech")
        if judge_hate is not None and bool(judge_hate) != current_hate:
            return {"state": "needs_review", "reason": "judge_disagreement"}

    return {"state": "auto_candidate", "reason": "legacy_current_agreement"}


def _legacy_summary(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "is_hate_speech": row.get("legacy_is_hate_speech"),
        "categories": row.get("legacy_categories", []),
        "fc": row.get("legacy_fc"),
        "text_hash": row.get("text_hash"),
    }


def _result_summary(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if row is None:
        return None
    payload = row.get("payload", row)
    return {
        "status": row.get("status", "succeeded"),
        "is_hate_speech": payload.get("is_hate_speech"),
        "categories": payload.get("categories", []),
        "variant": row.get("variant"),
        "repeat_index": row.get("repeat_index"),
    }


def _normal_categories(value: Any) -> set[str]:
    return {str(item) for item in _parse_categories(value)} - {"unclassified", "미분류", "혐오없음"}


def _parse_categories(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str):
        try:
            parsed = ast.literal_eval(value)
        except (SyntaxError, ValueError):
            return [value]
        if isinstance(parsed, list):
            return [str(item) for item in parsed]
    return [str(value)]


def _bool_or_none(value: Any) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    lowered = str(value).strip().lower()
    if lowered in {"true", "1", "yes", "y"}:
        return True
    if lowered in {"false", "0", "no", "n"}:
        return False
    return None


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _stable_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _write_jsonl(path: Path | str, rows: list[dict[str, Any]]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare semi-automatic gold construction assets from legacy result sets.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    convert_parser = subparsers.add_parser("convert", description="Convert a legacy workbook into experiment input and legacy label JSONL.")
    convert_parser.add_argument("--source", required=True, choices=[SOURCE_COMMENTS, SOURCE_SCRIPTS, SOURCE_INTERCODER])
    convert_parser.add_argument("--workbook-path", required=True)
    convert_parser.add_argument("--input-output-path", required=True)
    convert_parser.add_argument("--legacy-output-path", required=True)
    convert_parser.add_argument("--limit", type=int, default=None)

    queue_parser = subparsers.add_parser("queue", description="Build review queue from legacy labels and current RAG results.")
    queue_parser.add_argument("--legacy-label-path", required=True)
    queue_parser.add_argument("--current-results-path", required=True)
    queue_parser.add_argument("--output-path", required=True)
    queue_parser.add_argument("--judge-label-path", default=None)

    args = parser.parse_args()
    if args.command == "convert":
        input_count, legacy_count = convert_workbook(
            workbook_path=args.workbook_path,
            source=args.source,
            input_output_path=args.input_output_path,
            legacy_output_path=args.legacy_output_path,
            limit=args.limit,
        )
        print(f"wrote {input_count} inputs and {legacy_count} legacy labels")
    elif args.command == "queue":
        queue = build_review_queue(
            legacy_label_path=args.legacy_label_path,
            current_results_path=args.current_results_path,
            output_path=args.output_path,
            judge_label_path=args.judge_label_path,
        )
        states = {}
        for row in queue:
            states[row["review_state"]] = states.get(row["review_state"], 0) + 1
        print(" ".join(f"{state}={count}" for state, count in sorted(states.items())))


if __name__ == "__main__":
    main()
